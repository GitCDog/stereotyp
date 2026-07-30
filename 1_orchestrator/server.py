#!/usr/bin/env python3
"""
Lokaler Backend-Server für das Stereotypen Dashboard.

Verwendung:
  python server.py          # startet auf http://localhost:5000
"""

import csv
import json
import os
import subprocess
import sys
import threading
from datetime import datetime
from pathlib import Path

from flask import Flask, jsonify, request, send_file
from dotenv import load_dotenv

load_dotenv()

REPO_ROOT = Path(__file__).parent.parent
REIHENFOLGE_REL = "1_orchestrator/1_input/0_reihenfolge.txt"


TITLES_FILE = Path(__file__).parent / "generate_gpt_prompt_titled.py"


def generate_title(stereotyp: str, story_text: str) -> str:
    """Generiert provokanten Fragetitel via Claude CLI."""
    prompt = (
        f'Erstelle einen provokanten, viralen Fragetitel für diesen Instagram-Reel über das deutsche Stereotyp '
        f'"{stereotyp}". Der Titel soll: als Frage formuliert sein, max. 60 Zeichen, '
        f'neugierig machen, keine Anführungszeichen enthalten. '
        f'Antworte NUR mit dem Titel, nichts sonst.\n\nStory: {story_text[:300]}'
    )
    no_window = subprocess.CREATE_NO_WINDOW if hasattr(subprocess, "CREATE_NO_WINDOW") else 0
    env = {**os.environ, "TERM": "dumb", "NO_COLOR": "1"}
    result = subprocess.run(
        ["claude", "-p", prompt, "--allowedTools", ""],
        capture_output=True, text=True, encoding="utf-8",
        creationflags=no_window, timeout=30, env=env,
    )
    title = result.stdout.strip().strip('"').strip("'")
    return title if title else stereotyp


def save_title_to_dict(nr: int, title: str):
    """Fügt neuen Titel in TITLES-Dict von generate_gpt_prompt_titled.py ein."""
    content = TITLES_FILE.read_text(encoding="utf-8")
    if f"\n    {nr}:" in content:
        return  # bereits vorhanden
    escaped = title.replace('"', '\\"')
    new_entry = f'    {nr}: "{escaped}",\n'
    # Vor der schliessenden } des TITLES-Dicts einfügen
    content = content.replace("\n}\n\n\ndef _nr_str", f"\n{new_entry}}}\n\n\ndef _nr_str")
    TITLES_FILE.write_text(content, encoding="utf-8")


CSV_REL = "1_orchestrator/1_input/1_input_file.txt"


def _git_has_changes(*rel_paths: str) -> bool:
    no_window = subprocess.CREATE_NO_WINDOW
    for p in rel_paths:
        r = subprocess.run(
            ["git", "diff", "--quiet", p],
            cwd=str(REPO_ROOT), capture_output=True, creationflags=no_window,
        )
        if r.returncode != 0:
            return True
        r2 = subprocess.run(
            ["git", "diff", "--cached", "--quiet", p],
            cwd=str(REPO_ROOT), capture_output=True, creationflags=no_window,
        )
        if r2.returncode != 0:
            return True
    return False


def _git_commit_push(*rel_paths: str, message: str):
    no_window = subprocess.CREATE_NO_WINDOW
    for p in rel_paths:
        subprocess.run(["git", "add", p], cwd=str(REPO_ROOT),
                       check=True, capture_output=True, creationflags=no_window)
    subprocess.run(
        ["git", "commit", "-m", message],
        cwd=str(REPO_ROOT), check=True, capture_output=True, creationflags=no_window,
    )
    subprocess.run(["git", "push"], cwd=str(REPO_ROOT), check=True,
                   capture_output=True, creationflags=no_window)


def git_push_reihenfolge():
    """Committet und pusht 0_reihenfolge.txt nach GitHub."""
    try:
        if not _git_has_changes(REIHENFOLGE_REL):
            return
        _git_commit_push(REIHENFOLGE_REL, message="fix: Posting-Reihenfolge aktualisiert")
    except subprocess.CalledProcessError as e:
        print(f"[!] Git Push Reihenfolge fehlgeschlagen: {e}")


def git_push_csv():
    """Committet und pusht 1_input_file.txt nach GitHub, falls Änderungen vorhanden."""
    try:
        if not _git_has_changes(CSV_REL):
            return
        _git_commit_push(CSV_REL, message="fix: CSV nach Sync aktualisiert")
        print("[+] CSV-Änderungen auf GitHub gepusht")
    except subprocess.CalledProcessError as e:
        print(f"[!] Git Push CSV fehlgeschlagen: {e}")

app = Flask(__name__)

_server_start = datetime.now()

# Globaler Task-Status
_task = {"status": "idle", "message": "", "percent": 0, "log": []}
_task_lock = threading.Lock()

# Abort-Mechanismus
_current_proc = None
_proc_lock = threading.Lock()
_abort_flag = threading.Event()


def set_task(status, message, percent, log=None):
    with _task_lock:
        _task["status"] = status
        _task["message"] = message
        _task["percent"] = percent
        if log is not None:
            _task["log"] = log


def append_log(entry: str):
    with _task_lock:
        _task["log"].append(entry)


def parse_range(val: str) -> list[str]:
    """Unterstützt: '8', '4-10', '1,2,5'."""
    val = val.strip()
    if "," in val:
        return [v.strip() for v in val.split(",") if v.strip()]
    if "-" in val:
        parts = val.split("-")
        return [str(i) for i in range(int(parts[0]), int(parts[1]) + 1)]
    return [val]


def run_script(args: list[str]) -> int:
    global _current_proc
    if _abort_flag.is_set():
        return -1
    proc = subprocess.Popen(
        [sys.executable] + args,
        cwd=Path(__file__).parent,
        creationflags=subprocess.CREATE_NO_WINDOW,
    )
    with _proc_lock:
        _current_proc = proc
    try:
        proc.wait()
    finally:
        with _proc_lock:
            if _current_proc is proc:
                _current_proc = None
    return proc.returncode


def run_script_logged(args: list[str]) -> int:
    """Wie run_script, aber stdout/stderr wird live per append_log weitergegeben."""
    global _current_proc
    if _abort_flag.is_set():
        return -1
    proc = subprocess.Popen(
        [sys.executable] + args,
        cwd=Path(__file__).parent,
        stdout=subprocess.PIPE,
        stderr=subprocess.STDOUT,
        text=True,
        encoding="utf-8",
        errors="replace",
        creationflags=subprocess.CREATE_NO_WINDOW,
    )
    with _proc_lock:
        _current_proc = proc
    try:
        for line in proc.stdout:
            line = line.rstrip()
            if line:
                append_log(line)
        proc.wait()
    finally:
        with _proc_lock:
            if _current_proc is proc:
                _current_proc = None
    return proc.returncode


def refresh_dashboard():
    run_script(["sync_status.py"])
    run_script(["generate_dashboard.py"])


# ── Seiten ──────────────────────────────────────────────────────────────────

@app.route("/")
def index():
    from flask import make_response
    resp = make_response(send_file("dashboard.html"))
    resp.headers["Cache-Control"] = "no-cache, no-store, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


# ── Server-Status ────────────────────────────────────────────────────────────

@app.route("/api/status")
def status():
    uptime = datetime.now() - _server_start
    h, rem = divmod(int(uptime.total_seconds()), 3600)
    m, s = divmod(rem, 60)
    return jsonify({
        "started": _server_start.strftime("%d.%m.%Y %H:%M:%S"),
        "uptime": f"{h:02d}:{m:02d}:{s:02d}",
        "online": True,
    })


# ── Fortschritt ─────────────────────────────────────────────────────────────

@app.route("/api/progress")
def progress():
    with _task_lock:
        return jsonify(dict(_task))


# ── Story-Text ───────────────────────────────────────────────────────────────

@app.route("/api/story-text")
def story_text():
    nr = request.args.get("nr", "").strip()
    if not nr:
        return jsonify({"error": "nr fehlt"}), 400
    try:
        nr4 = f"{int(nr):04d}"
    except ValueError:
        return jsonify({"error": "ungültige nr"}), 400
    stories_dir = Path(__file__).parent / "1_input"
    matches = [p for p in stories_dir.glob(f"{nr4}_*.txt")
               if p.name not in ("00_sammelsurium.txt",)]
    if not matches:
        return jsonify({"text": None})
    text = matches[0].read_text(encoding="utf-8").strip()
    return jsonify({"text": text})


# ── Audio ────────────────────────────────────────────────────────────────────

@app.route("/api/generate-audio", methods=["POST"])
def generate_audio():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    body = request.get_json(silent=True) or {}
    story_val = str(body.get("story", "")).strip()

    def task():
        try:
            if story_val:
                numbers = parse_range(story_val)
                candidates = [(nr, "") for nr in numbers]
            else:
                input_file = Path(__file__).parent / "1_input" / "1_input_file.txt"
                with open(input_file, encoding="utf-8") as f:
                    rows = list(csv.DictReader(f))
                candidates = [
                    (r["nr"].strip(), r.get("stereotyp", "").strip()) for r in rows
                    if r.get("status_story") == "X"
                    and r.get("status_audio", "") != "X"
                ]

            total = len(candidates)
            if not total:
                set_task("complete", "Keine ausstehenden Audios.", 100, log=[])
                return

            log = [f"⏳ #{nr}  {name}" for nr, name in candidates]
            set_task("running", f"0/{total} fertig", 5, log=log)

            for i, (nr, name) in enumerate(candidates):
                pct = int((i / total) * 90)
                log[i] = f"🔄 #{nr}  {name}"
                set_task("running", f"{i}/{total} fertig – generiere #{nr}...", pct, log=list(log))
                run_script(["generate_audio.py", "--story", str(nr)])
                log[i] = f"✅ #{nr}  {name}"
                set_task("running", f"{i+1}/{total} fertig", int(((i+1) / total) * 90), log=list(log))

            set_task("running", "Dashboard aktualisieren...", 95, log=list(log))
            refresh_dashboard()
            set_task("complete", f"Fertig! {total} Audio(s) generiert.", 100, log=list(log))
        except Exception as e:
            set_task("error", str(e), 0)

    set_task("running", "Starte...", 5, log=[])
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Audio für alle Pics ──────────────────────────────────────────────────────

@app.route("/api/generate-audio-for-pics", methods=["POST"])
def generate_audio_for_pics():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    def task():
        try:
            import csv
            input_file = Path(__file__).parent / "1_input" / "1_input_file.txt"
            with open(input_file, encoding="utf-8") as f:
                rows = list(csv.DictReader(f))
            candidates = [
                (r["nr"].strip(), r.get("stereotyp", "").strip()) for r in rows
                if r.get("status_pic") == "X"
                and r.get("status_audio", "") != "X"
                and r.get("status_story", "") == "X"
            ]
            total = len(candidates)
            if not total:
                set_task("complete", "Keine ausstehenden Audios für vorhandene Bilder.", 100, log=[])
                return

            log = [f"⏳ #{nr}  {name}" for nr, name in candidates]
            set_task("running", f"0/{total} fertig", 5, log=log)

            for i, (nr, name) in enumerate(candidates):
                pct = int(((i) / total) * 90)
                log[i] = f"🔄 #{nr}  {name}"
                set_task("running", f"{i}/{total} fertig – generiere #{nr}...", pct, log=list(log))
                run_script(["generate_audio.py", "--story", nr])
                log[i] = f"✅ #{nr}  {name}"
                set_task("running", f"{i+1}/{total} fertig", int(((i+1) / total) * 90), log=list(log))

            set_task("running", "Dashboard aktualisieren...", 95, log=list(log))
            refresh_dashboard()
            set_task("complete", f"Fertig! {total} Audio(s) generiert.", 100, log=list(log))
        except Exception as e:
            set_task("error", str(e), 0)

    set_task("running", "Starte...", 5, log=[])
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Bild ─────────────────────────────────────────────────────────────────────

@app.route("/api/generate-picture", methods=["POST"])
def generate_picture():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    body = request.get_json(silent=True) or {}
    story_val = str(body.get("story", "")).strip()

    def task():
        try:
            if story_val:
                set_task("running", f"Bild für {story_val}...", 10)
                run_script(["generate_pictures.py", story_val])
            set_task("running", "Dashboard aktualisieren...", 95)
            refresh_dashboard()
            set_task("complete", "Fertig!", 100)
        except Exception as e:
            set_task("error", str(e), 0)

    set_task("running", "Starte...", 5)
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Neue Story erstellen ─────────────────────────────────────────────────────

@app.route("/api/create-story", methods=["POST"])
def create_story():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    body = request.get_json(silent=True) or {}
    stereotyp = body.get("stereotyp", "").strip()
    stichworte = body.get("stichworte", "").strip()
    position = body.get("position", "")  # "start", "end", Zahl, oder leer = nicht einreihen

    if not stereotyp:
        return jsonify({"error": "Stereotyp fehlt"}), 400

    def task():
        try:
            input_file = Path(__file__).parent / "1_input" / "1_input_file.txt"
            with open(input_file, encoding="utf-8") as f:
                rows = list(csv.DictReader(f))

            # Prüfen ob Stereotyp schon existiert
            existing = next((r for r in rows if r.get("stereotyp", "").strip().lower() == stereotyp.lower()), None)
            if existing:
                new_nr = existing["nr"].strip()
                log = [f"ℹ️ #{new_nr} '{stereotyp}' existiert bereits – Story wird neu generiert"]
                set_task("running", f"Regeneriere Story #{new_nr}...", 10, log=log)
            else:
                max_nr = max(int(r["nr"].strip()) for r in rows if r.get("nr", "").strip().isdigit())
                new_nr = str(max_nr + 1)
                fieldnames = list(rows[0].keys())
                new_row = {k: "" for k in fieldnames}
                new_row["nr"] = new_nr
                new_row["stereotyp"] = stereotyp
                with open(input_file, "a", encoding="utf-8", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writerow(new_row)
                log = [f"📋 #{new_nr} '{stereotyp}' in CSV eingetragen"]
                if stichworte:
                    log.append(f"🔑 Stichworte: {stichworte}")
                set_task("running", f"Story #{new_nr} generieren...", 10, log=log)

            log.append(f"⏳ Story-Text via Claude generieren...")
            set_task("running", f"Story #{new_nr} via Claude generieren...", 30, log=list(log))
            args = ["generate_stories.py", "--story", new_nr]
            if stichworte:
                args += ["--stichworte", stichworte]
            run_script(args)
            log.append(f"✅ Story-Text gespeichert")

            # Provokanten Titel generieren und in TITLES-Dict speichern
            set_task("running", f"Titel für #{new_nr} generieren...", 45, log=list(log))
            try:
                story_file = next(
                    (p for p in (Path(__file__).parent / "1_input").glob(f"{int(new_nr):04d}_*.txt")
                     if p.name not in {"00_sammelsurium.txt", "gpt_prompts.txt"}),
                    None
                )
                story_text = story_file.read_text(encoding="utf-8").strip() if story_file else stereotyp
                title = generate_title(stereotyp, story_text)
                save_title_to_dict(int(new_nr), title)
                log.append(f"🏷 Titel: \"{title}\"")
            except Exception as e:
                log.append(f"⚠️ Titel-Generierung fehlgeschlagen: {e}")

            log.append(f"⏳ Caption via Claude generieren...")
            set_task("running", f"Caption für #{new_nr} generieren...", 65, log=list(log))
            caption_args = ["generate_captions.py", "--story", new_nr]
            if stichworte:
                caption_args += ["--stichworte", stichworte]
            run_script(caption_args)
            log.append(f"✅ Caption fertig")

            if position:
                reihenfolge_path = Path(__file__).parent / "1_input" / "0_reihenfolge.txt"
                lines = []
                if reihenfolge_path.exists():
                    lines = [l.strip() for l in reihenfolge_path.read_text(encoding="utf-8").splitlines() if l.strip()]
                lines = [l for l in lines if l != new_nr]
                if position == "start":
                    lines.insert(0, new_nr)
                    log.append(f"📋 #{new_nr} an den Anfang der Reihenfolge gesetzt")
                elif position == "end":
                    lines.append(new_nr)
                    log.append(f"📋 #{new_nr} ans Ende der Reihenfolge gesetzt")
                else:
                    try:
                        pos = max(0, int(position) - 1)
                        lines.insert(pos, new_nr)
                        log.append(f"📋 #{new_nr} an Position {position} der Reihenfolge gesetzt")
                    except Exception:
                        lines.append(new_nr)
                reihenfolge_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
                git_push_reihenfolge()

            set_task("running", "GPT-Prompt aktualisieren...", 88, log=list(log))
            onedrive_out = r"C:\Users\slawa\OneDrive\8_stereotypen\gpt_prompts.txt"
            run_script(["generate_gpt_prompt_titled.py"])
            log.append(f"📝 GPT-Prompt aktualisiert")

            log.append(f"📷 Bild muss noch manuell erstellt werden (GPT)")
            set_task("running", "Dashboard aktualisieren...", 92, log=list(log))
            refresh_dashboard()
            set_task("complete", f"Story #{new_nr} '{stereotyp}' erstellt + vertont!", 100, log=list(log))
        except Exception as e:
            set_task("error", str(e), 0)

    set_task("running", "Starte...", 5, log=[])
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Story ────────────────────────────────────────────────────────────────────

@app.route("/api/generate-story", methods=["POST"])
def generate_story():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    body = request.get_json(silent=True) or {}
    story_val = str(body.get("story", "")).strip()

    def task():
        try:
            if story_val:
                numbers = parse_range(story_val)
                total = len(numbers)
                for i, nr in enumerate(numbers):
                    set_task("running", f"Story #{nr}...", int((i / total) * 90))
                    run_script(["generate_stories.py", "--story", str(nr)])
            else:
                set_task("running", "Nächste ausstehende Story...", 20)
                run_script(["generate_stories.py"])
            set_task("running", "Dashboard aktualisieren...", 95)
            refresh_dashboard()
            set_task("complete", "Fertig!", 100)
        except Exception as e:
            set_task("error", str(e), 0)

    set_task("running", "Starte...", 5)
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Caption ──────────────────────────────────────────────────────────────────

@app.route("/api/generate-caption", methods=["POST"])
def generate_caption():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    body = request.get_json(silent=True) or {}
    story_val = str(body.get("story", "")).strip()

    def task():
        try:
            if story_val:
                numbers = parse_range(story_val)
                total = len(numbers)
                for i, nr in enumerate(numbers):
                    set_task("running", f"Caption #{nr}...", int((i / total) * 90))
                    run_script(["generate_captions.py", "--story", str(nr)])
            else:
                set_task("running", "Alle ausstehenden Captions...", 10)
                run_script(["generate_captions.py"])
            set_task("running", "Dashboard aktualisieren...", 95)
            refresh_dashboard()
            set_task("complete", "Fertig!", 100)
        except Exception as e:
            set_task("error", str(e), 0)

    set_task("running", "Starte...", 5)
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Video ────────────────────────────────────────────────────────────────────

@app.route("/api/generate-video", methods=["POST"])
def generate_video():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    body = request.get_json(silent=True) or {}
    story_val = str(body.get("story", "")).strip()

    def task():
        _abort_flag.clear()
        try:
            input_file = Path(__file__).parent / "1_input" / "1_input_file.txt"
            if story_val:
                numbers = parse_range(story_val)
                with open(input_file, encoding="utf-8") as f:
                    all_rows = {r["nr"].strip(): r.get("stereotyp", "").strip()
                                for r in csv.DictReader(f)}
                candidates = [(nr, all_rows.get(nr, "")) for nr in numbers]
            else:
                with open(input_file, encoding="utf-8") as f:
                    rows = list(csv.DictReader(f))
                candidates = [
                    (r["nr"].strip(), r.get("stereotyp", "").strip()) for r in rows
                    if r.get("status_audio") == "X"
                    and r.get("status_pic") == "X"
                    and r.get("status_video", "") != "X"
                ]

            total = len(candidates)
            if not total:
                set_task("complete", "Keine Videos ausstehend.", 100, log=[])
                return

            log = [f"⏳ #{nr}  {name}" for nr, name in candidates]
            set_task("running", f"0/{total} Videos fertig", 5, log=log)

            for i, (nr, name) in enumerate(candidates):
                if _abort_flag.is_set():
                    set_task("idle", "Abgebrochen", 0, log=list(log))
                    return
                pct = int((i / total) * 88) + 5
                log[i] = f"🎬 #{nr}  {name}"
                set_task("running", f"{i}/{total} – rendere #{nr}: {name}...", pct, log=list(log))
                run_script(["generate_videos.py", "--story", str(nr), "--subtitles", "--bg-music", "Demeter - O.P..mp3", "--bg-music-volume", "0.2"])
                if _abort_flag.is_set():
                    log[i] = f"⏹️ #{nr}  {name}"
                    set_task("idle", "Abgebrochen", 0, log=list(log))
                    return
                log[i] = f"✅ #{nr}  {name}"
                set_task("running", f"{i+1}/{total} fertig", int(((i + 1) / total) * 88) + 5, log=list(log))

            set_task("running", "Dashboard aktualisieren...", 97, log=list(log))
            refresh_dashboard()
            set_task("complete", f"Fertig! {total} Video(s) erstellt.", 100, log=list(log))
        except Exception as e:
            set_task("error", str(e), 0)

    _abort_flag.clear()
    set_task("running", "Starte...", 5, log=[])
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Bilder via Playwright (ChatGPT) ─────────────────────────────────────────

@app.route("/api/generate-pictures-playwright", methods=["POST"])
def generate_pictures_playwright():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    body = request.get_json(silent=True) or {}
    story_val = str(body.get("story", "")).strip()

    def task():
        _abort_flag.clear()
        try:
            target = story_val
            if not target:
                import csv as _csv
                input_file = Path(__file__).parent / "1_input" / "1_input_file.txt"
                with open(input_file, encoding="utf-8") as f:
                    rows = list(_csv.DictReader(f))
                pending = [r["nr"] for r in rows
                           if r.get("status_story") == "X" and r.get("status_pic") != "X"]
                if not pending:
                    set_task("complete", "Alle Stories haben bereits ein Bild.", 100)
                    return
                target = ",".join(pending[:5])
                append_log(f"[*] Auto-Modus: {len(pending)} Stories ohne Bild → nächste 5: {target}")
            set_task("running", f"Playwright: Bilder für {target}...", 10, log=[])
            run_script_logged(["generate_pictures_playwright.py", "--story", target])
            set_task("running", "Dashboard aktualisieren...", 95)
            refresh_dashboard()
            set_task("complete", f"Fertig! Bilder für {story_val} generiert.", 100)
        except Exception as e:
            set_task("error", str(e), 0)

    _abort_flag.clear()
    set_task("running", "Starte Playwright...", 5)
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Instagram Post ───────────────────────────────────────────────────────────

@app.route("/api/instagram-post", methods=["POST"])
def instagram_post():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    body = request.get_json(silent=True) or {}
    story_val = str(body.get("story", "")).strip()

    def task():
        try:
            if story_val:
                # Manueller Post: via GitHub Actions triggern + auf Ergebnis warten
                import time as _time
                numbers = parse_range(story_val)
                repo_root = Path(__file__).parent.parent
                gh_env = {k: v for k, v in os.environ.items()
                          if k not in ("GITHUB_TOKEN", "GH_TOKEN")}
                input_file = Path(__file__).parent / "1_input" / "1_input_file.txt"

                final_notif = None
                for i, nr in enumerate(numbers):
                    # Story-Name aus CSV lesen
                    stereotyp_name = f"#{nr}"
                    try:
                        with open(input_file, encoding="utf-8") as f:
                            for row in csv.DictReader(f):
                                if str(row.get("nr", "")).strip() == str(nr).strip():
                                    stereotyp_name = row.get("stereotyp", f"#{nr}").strip()
                                    break
                    except Exception:
                        pass

                    append_log(f"{'─'*50}")
                    append_log(f"▶ Starte GitHub Actions für #{nr} '{stereotyp_name}'")
                    set_task("running", f"Triggere Workflow fuer '{stereotyp_name}'...", 10)

                    result = subprocess.run(
                        ["gh", "workflow", "run", "post_story.yml", "-f", f"story_nr={nr}"],
                        cwd=str(repo_root),
                        capture_output=True, text=True, encoding="utf-8", errors="replace",
                        timeout=30, env=gh_env,
                    )
                    if result.returncode != 0:
                        err = result.stderr.strip() or result.stdout.strip()
                        append_log(f"❌ Workflow-Start fehlgeschlagen: {err}")
                        final_notif = {"type": "error", "message": f"Workflow-Start fehlgeschlagen:\n{err}"}
                        break

                    run_url = result.stdout.strip()
                    run_id = run_url.rstrip("/").split("/")[-1]
                    append_log(f"⏳ Warte auf Ergebnis (Run #{run_id})...")

                    # Polling: max 7 Minuten, alle 10 Sekunden
                    conclusion = None
                    for attempt in range(42):
                        _time.sleep(10)
                        pr = subprocess.run(
                            ["gh", "run", "view", run_id, "--json", "status,conclusion"],
                            cwd=str(repo_root), capture_output=True, text=True,
                            encoding="utf-8", errors="replace", env=gh_env,
                        )
                        if pr.returncode == 0:
                            d = json.loads(pr.stdout)
                            if d.get("status") == "completed":
                                conclusion = d.get("conclusion")
                                break
                        pct = 10 + int(((attempt + 1) / 42) * 80)
                        set_task("running", f"⏳ Warte auf GitHub Actions... ({(attempt+1)*10}s)", pct)

                    if conclusion == "success":
                        append_log(f"✅ '{stereotyp_name}' erfolgreich gepostet!")
                        final_notif = {"type": "success", "message": f"'{stereotyp_name}' wurde erfolgreich gepostet!"}
                    elif conclusion:
                        fl = subprocess.run(
                            ["gh", "run", "view", run_id, "--log-failed"],
                            cwd=str(repo_root), capture_output=True, text=True,
                            encoding="utf-8", errors="replace", env=gh_env,
                        )
                        error_lines = [l for l in fl.stdout.splitlines()
                                       if "error" in l.lower() or "Error" in l or "[-]" in l][-5:]
                        detail = "\n".join(error_lines) or conclusion
                        append_log(f"❌ Fehlgeschlagen ({conclusion}): {detail}")
                        final_notif = {"type": "error", "message": f"Posting fehlgeschlagen:\n{detail}"}
                    else:
                        append_log(f"⏰ Timeout: kein Ergebnis nach 7 Minuten")
                        final_notif = {"type": "error", "message": "Timeout: kein Ergebnis nach 7 Minuten"}

                with _task_lock:
                    _task["reload"] = False
                    _task["notification"] = final_notif
                status = "complete" if (final_notif or {}).get("type") == "success" else "error"
                set_task(status, final_notif["message"] if final_notif else "Fertig", 100)
            else:
                # Auto-Modus: lokaler instagram_poster (nächste ausstehende Story)
                set_task("running", "Poste nächste Story lokal...", 5, log=[])
                env = os.environ.copy()
                env["FORCE_POST"] = "1"
                proc = subprocess.Popen(
                    [sys.executable, "instagram_poster.py"],
                    cwd=Path(__file__).parent,
                    env=env,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                    encoding="utf-8",
                    errors="replace",
                    creationflags=subprocess.CREATE_NO_WINDOW,
                )
                for line in proc.stdout:
                    line = line.rstrip()
                    if line:
                        append_log(line)
                proc.wait()
                if proc.returncode != 0:
                    append_log(f"⚠️ Fehlercode {proc.returncode}")
                set_task("running", "Dashboard aktualisieren...", 97)
                refresh_dashboard()
                set_task("complete", "Fertig!", 100)
        except Exception as e:
            set_task("error", str(e), 0)

    set_task("running", "Starte...", 5, log=[])
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Refresh (Datei-Scan + Dashboard + neue Stories) ─────────────────────────

@app.route("/api/refresh", methods=["POST"])
def refresh():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    def task():
        _abort_flag.clear()
        log = []
        try:
            set_task("running", "GitHub: aktuelle Änderungen holen (git pull)...", 5, log=list(log))
            repo_root = Path(__file__).parent.parent
            pull = subprocess.run(
                ["git", "pull", "--rebase"],
                cwd=str(repo_root), capture_output=True, text=True, encoding="utf-8", errors="replace",
            )
            if "Already up to date" in pull.stdout:
                log.append("✅ GitHub: Bereits aktuell")
            else:
                log.append(f"📥 GitHub pull: {pull.stdout.strip().splitlines()[-1]}")

            onedrive_dir = Path(r"C:\Users\slawa\OneDrive\8_stereotypen")
            def _count_imgs(d): return sum(len(list(d.glob(f"*.{e}"))) for e in ("png","jpg","jpeg")) if d.exists() else 0
            imgs_before = _count_imgs(onedrive_dir)

            set_task("running", "OneDrive: neue Bilder identifizieren...", 15, log=list(log))
            run_script(["onedrive_check.py", "--onedrive-only"])

            imgs_after = _count_imgs(onedrive_dir)
            processed = imgs_before - imgs_after
            if processed > 0:
                log.append(f"🖼️ {processed} Bild(er) aus OneDrive identifiziert und verarbeitet")
            else:
                log.append("✅ OneDrive: Keine neuen Bilder")

            set_task("running", "Sammelsurium → 1_input/ extrahieren...", 35, log=list(log))
            import sync_status as ss
            new_from_sammelsurium = ss.check_sammelsurium()
            if new_from_sammelsurium:
                log.append(f"📄 {new_from_sammelsurium} neue Stories aus Sammelsurium extrahiert")
            else:
                log.append("✅ Sammelsurium: Keine neuen Stories")

            set_task("running", "Dateien scannen & CSV aktualisieren...", 65, log=list(log))
            run_script(["sync_status.py"])
            log.append("🔄 CSV und Dateistatus synchronisiert")

            set_task("running", "GPT-Prompts für fehlende Bilder schreiben...", 80, log=list(log))
            onedrive_out = r"C:\Users\slawa\OneDrive\8_stereotypen\gpt_prompts.txt"
            run_script(["generate_gpt_prompt_titled.py"])
            log.append("📝 GPT Prompts aktualisiert")

            set_task("running", "Dashboard aktualisieren...", 90, log=list(log))
            refresh_dashboard()
            log.append("✅ Dashboard aktualisiert")

            set_task("running", "CSV nach GitHub pushen...", 95, log=list(log))
            git_push_csv()
            log.append("📤 CSV auf GitHub synchronisiert")

            msg = f"Fertig! {new_from_sammelsurium} neue Stories extrahiert." if new_from_sammelsurium else "Fertig!"
            set_task("complete", msg, 100, log=list(log))
        except Exception as e:
            set_task("error", str(e), 0)

    _abort_flag.clear()
    set_task("running", "Starte Refresh...", 5, log=[])
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


# ── Restart ──────────────────────────────────────────────────────────────────

@app.route("/api/restart", methods=["POST"])
def restart_server():
    def do_restart():
        import time, os
        time.sleep(1)
        subprocess.Popen(
            [sys.executable, "server.py"],
            cwd=Path(__file__).parent,
            creationflags=subprocess.DETACHED_PROCESS | subprocess.CREATE_NEW_PROCESS_GROUP,
        )
        time.sleep(0.5)
        os._exit(0)
    threading.Thread(target=do_restart, daemon=True).start()
    return jsonify({"status": "restarting"})


# ── Abort ────────────────────────────────────────────────────────────────────

@app.route("/api/abort", methods=["POST"])
def abort_task():
    _abort_flag.set()
    with _proc_lock:
        proc = _current_proc
    if proc and proc.poll() is None:
        proc.terminate()
    with _task_lock:
        _task["status"] = "idle"
        _task["message"] = "Abgebrochen"
        _task["percent"] = 0
    return jsonify({"status": "aborted"})


# ── Reihenfolge ──────────────────────────────────────────────────────────────

def _video_on_cloudinary(nr) -> bool:
    """Prüft via CSV ob Video auf Cloudinary liegt.
    Bedingung: status_video=X und insta_post leer (noch nicht gepostet/gelöscht)."""
    input_file = Path(__file__).parent / "1_input" / "1_input_file.txt"
    nr_str = str(int(str(nr).strip()))
    try:
        with open(input_file, encoding="utf-8") as f:
            for row in csv.DictReader(f):
                if str(row.get("nr", "")).strip() == nr_str:
                    return row.get("status_video", "").strip() == "X" and not row.get("insta_post", "").strip()
    except Exception:
        pass
    return False


@app.route("/api/cloudinary-check")
def cloudinary_check():
    nr = request.args.get("nr", "").strip()
    if not nr:
        return jsonify({"error": "nr fehlt"}), 400
    exists = _video_on_cloudinary(nr)
    return jsonify({"exists": exists, "nr": nr})


@app.route("/api/reihenfolge", methods=["GET"])
def get_reihenfolge():
    path = Path(__file__).parent / "1_input" / "0_reihenfolge.txt"
    lines = []
    if path.exists():
        lines = [l.strip() for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]
    return jsonify({"reihenfolge": lines})


@app.route("/api/reihenfolge", methods=["POST"])
def set_reihenfolge():
    body = request.get_json(silent=True) or {}
    nr = str(body.get("nr", "")).strip()
    position = body.get("position", "end")  # "start", "end", oder Zahl

    if not nr:
        return jsonify({"error": "nr fehlt"}), 400

    try: nr = str(int(nr))
    except ValueError: pass

    path = Path(__file__).parent / "1_input" / "0_reihenfolge.txt"
    lines = []
    if path.exists():
        lines = [l.strip() for l in path.read_text(encoding="utf-8").splitlines() if l.strip()]

    lines = [l for l in lines if l != nr]  # doppelte entfernen

    if position == "start":
        lines.insert(0, nr)
    elif position == "end":
        lines.append(nr)
    else:
        try:
            pos = max(0, int(position) - 1)
            lines.insert(pos, nr)
        except Exception:
            lines.append(nr)

    path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    threading.Thread(target=git_push_reihenfolge, daemon=True).start()
    return jsonify({"status": "ok", "reihenfolge": lines})


@app.route("/api/reihenfolge/remove", methods=["POST"])
def remove_reihenfolge():
    body = request.get_json(silent=True) or {}
    nr = str(body.get("nr", "")).strip()
    if not nr:
        return jsonify({"error": "nr fehlt"}), 400

    path = Path(__file__).parent / "1_input" / "0_reihenfolge.txt"
    if path.exists():
        def _norm(s):
            try: return str(int(s))
            except: return s
        nr_norm = _norm(nr)
        lines = [l.strip() for l in path.read_text(encoding="utf-8").splitlines()
                 if l.strip() and _norm(l.strip()) != nr_norm]
        path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    threading.Thread(target=git_push_reihenfolge, daemon=True).start()
    return jsonify({"status": "ok"})


# ── Mark Posted ──────────────────────────────────────────────────────────────

@app.route("/api/mark-posted", methods=["POST"])
def mark_posted():
    body = request.get_json(silent=True) or {}
    nr = body.get("nr")
    if not nr:
        return jsonify({"error": "nr fehlt"}), 400
    import input_reader as ir
    ir.update_field(int(nr), "insta_post", "X", "1_input/1_input_file.txt")
    refresh_dashboard()
    return jsonify({"status": "ok"})


@app.route("/api/unmark-posted", methods=["POST"])
def unmark_posted():
    body = request.get_json(silent=True) or {}
    nr = body.get("nr")
    if not nr:
        return jsonify({"error": "nr fehlt"}), 400
    import input_reader as ir
    ir.update_field(int(nr), "insta_post", "", "1_input/1_input_file.txt")
    refresh_dashboard()
    return jsonify({"status": "ok"})


@app.route("/api/reset-video", methods=["POST"])
def reset_video():
    """Setzt eine Story zurück damit das Video neu generiert werden kann.
    Löscht MP4, verschiebt Audio+Bild zurück nach output/, resettet Status-Felder."""
    body = request.get_json(silent=True) or {}
    nr = body.get("nr")
    if not nr:
        return jsonify({"error": "nr fehlt"}), 400

    import input_reader as ir
    import shutil

    nr_str = f"{int(str(nr).strip()):04d}"
    output_dir = Path("output")
    used_dir = Path("output/0_used")
    msgs = []

    # MP4 löschen
    for mp4 in list(output_dir.glob(f"{nr_str}*.mp4")) + list(used_dir.glob(f"{nr_str}*.mp4")):
        mp4.unlink()
        msgs.append(f"🗑 {mp4.name} gelöscht")

    # Audio zurück nach output/ (bleibt erhalten, status_audio=X bleibt)
    for mp3 in used_dir.glob(f"{nr_str}*.mp3"):
        dst = output_dir / mp3.name
        if not dst.exists():
            shutil.move(str(mp3), str(dst))
        msgs.append(f"Audio zurueck: {mp3.name}")

    # Bild zurück nach output/ (NIEMALS löschen)
    for pic in list(used_dir.glob(f"{nr_str}_pic*.*")):
        dst = output_dir / pic.name
        if not dst.exists():
            shutil.move(str(pic), str(dst))
        msgs.append(f"Bild zurueck: {pic.name}")

    # CSV zurücksetzen
    f = "1_input/1_input_file.txt"
    for field in ("status_pic", "status_video", "insta_post", "youtube_post"):
        ir.update_field(str(nr), field, "", f)
    msgs.append("✅ status_pic, status_video, insta_post, youtube_post zurückgesetzt")

    refresh_dashboard()
    return jsonify({"status": "ok", "messages": msgs})


# ── Sofort posten (Vollautomatisch aus story_jetzt.xlsx) ─────────────────────

STORY_JETZT_XLSX = Path(r"C:\Users\slawa\OneDrive\8_stereotypen\story_jetzt.xlsx")


@app.route("/api/sofort-posten", methods=["POST"])
def sofort_posten():
    if _task["status"] == "running":
        return jsonify({"error": "Task läuft bereits"}), 409

    def task():
        try:
            import openpyxl
            import input_reader as ir
            import traceback

            if not STORY_JETZT_XLSX.exists():
                set_task("error", f"Datei nicht gefunden: {STORY_JETZT_XLSX}", 0)
                return

            wb = openpyxl.load_workbook(str(STORY_JETZT_XLSX))
            ws = wb.active
            headers = [cell.value for cell in ws[1]]

            def col(name):
                return headers.index(name) + 1 if name in headers else None

            stereotyp_col = col("stereotyp")
            stichworte_col = col("stichworte")
            titel_col = col("Titel")
            status_col = col("status")

            if stereotyp_col is None or status_col is None:
                set_task("error", "story_jetzt.xlsx: Spalten 'stereotyp' und 'status' fehlen", 0)
                return

            pending_row_idx = None
            for i, row in enumerate(ws.iter_rows(min_row=2, values_only=False), start=2):
                if not row[status_col - 1].value:
                    pending_row_idx = i
                    break

            if pending_row_idx is None:
                set_task("complete", "Keine ausstehenden Einträge in story_jetzt.xlsx", 100,
                         log=["✅ story_jetzt.xlsx: Alle Einträge erledigt"])
                return

            stereotyp = (ws.cell(pending_row_idx, stereotyp_col).value or "").strip()
            stichworte = (ws.cell(pending_row_idx, stichworte_col).value or "").strip() if stichworte_col else ""
            xlsx_titel = (ws.cell(pending_row_idx, titel_col).value or "").strip() if titel_col else ""

            if not stereotyp:
                set_task("error", f"Zeile {pending_row_idx}: Stereotyp ist leer", 0)
                return

            set_task("running", f"Starte Pipeline für: {stereotyp}", 5, log=[
                f"📋 Stereotyp: {stereotyp}",
                *([ f"🔑 Stichworte: {stichworte}"] if stichworte else []),
                "─" * 40,
            ])

            # 1. In CSV eintragen
            input_file = Path(__file__).parent / "1_input" / "1_input_file.txt"
            with open(input_file, encoding="utf-8") as f:
                rows = list(csv.DictReader(f))

            existing = next((r for r in rows if r.get("stereotyp", "").strip().lower() == stereotyp.lower()), None)
            if existing:
                new_nr = existing["nr"].strip()
                append_log(f"ℹ️  [1/8] CSV: #{new_nr} bereits vorhanden")
            else:
                max_nr = max(int(r["nr"].strip()) for r in rows if r.get("nr", "").strip().isdigit())
                new_nr = str(max_nr + 1)
                fieldnames = list(rows[0].keys())
                new_row = {k: "" for k in fieldnames}
                new_row["nr"] = new_nr
                new_row["stereotyp"] = stereotyp
                with open(input_file, "a", encoding="utf-8", newline="") as f:
                    writer = csv.DictWriter(f, fieldnames=fieldnames)
                    writer.writerow(new_row)
                append_log(f"✅ [1/8] CSV: #{new_nr} eingetragen")

            # 2. Story generieren
            set_task("running", f"[2/8] Story #{new_nr} via Claude...", 12)
            append_log(f"⏳ [2/8] Story via Claude generieren...")
            args = ["generate_stories.py", "--story", new_nr]
            if stichworte:
                args += ["--stichworte", stichworte]
            run_script_logged(args)
            append_log(f"✅ [2/8] Story-Text gespeichert")

            # 3. Titel – Excel-Titel hat Vorrang, dann TITLES-Dict, sonst neu generieren
            set_task("running", f"[3/8] Titel für #{new_nr}...", 20)
            append_log(f"⏳ [3/8] Provokanten Titel generieren...")
            try:
                from generate_gpt_prompt_titled import TITLES as _titles
                if xlsx_titel:
                    # Titel aus Excel übernehmen
                    save_title_to_dict(int(new_nr), xlsx_titel)
                    append_log(f"✅ [3/8] Titel aus Excel: \"{xlsx_titel}\"")
                elif int(new_nr) in _titles:
                    append_log(f"✅ [3/8] Titel bereits vorhanden: \"{_titles[int(new_nr)]}\"")
                else:
                    story_file = next(
                        (p for p in (Path(__file__).parent / "1_input").glob(f"{int(new_nr):04d}_*.txt")
                         if p.name not in {"00_sammelsurium.txt", "gpt_prompts.txt"}),
                        None
                    )
                    story_text_content = story_file.read_text(encoding="utf-8").strip() if story_file else stereotyp
                    title = generate_title(stereotyp, story_text_content)
                    save_title_to_dict(int(new_nr), title)
                    append_log(f"✅ [3/8] Titel generiert: \"{title}\"")
            except Exception as e:
                append_log(f"⚠️  [3/8] Titel fehlgeschlagen: {e}")

            # 4. Caption generieren
            set_task("running", f"[4/8] Caption #{new_nr}...", 27)
            append_log(f"⏳ [4/8] Caption via Claude generieren...")
            caption_args = ["generate_captions.py", "--story", new_nr]
            if stichworte:
                caption_args += ["--stichworte", stichworte]
            run_script_logged(caption_args)
            append_log(f"✅ [4/8] Caption fertig")

            # 5. Bild generieren (Playwright / ChatGPT)
            _row_check = ir.find_row(new_nr, input_file)
            if _row_check and _row_check.get("status_pic") == "X":
                append_log(f"✅ [5/8] Bild bereits vorhanden – überspringe")
            else:
                set_task("running", f"[5/8] Bild #{new_nr} via Playwright – kann Minuten dauern...", 35)
                append_log(f"⏳ [5/8] Bild via Playwright (ChatGPT) generieren...")
                run_script_logged(["generate_pictures_playwright.py", "--story", new_nr])
                ns = f"{int(new_nr):04d}"
                out_dir = Path(__file__).parent / "output"
                pic_path_found = any(out_dir.glob(f"{ns}_pic*.png")) or any(out_dir.glob(f"{ns}_pic*.jpg"))
                pic_used_found = any((out_dir / "0_used").glob(f"{ns}_pic*.png")) or any((out_dir / "0_used").glob(f"{ns}_pic*.jpg"))
                if not pic_path_found and not pic_used_found:
                    append_log("❌ [5/8] Kein Bild erstellt – Pipeline abgebrochen (ChatGPT-Limit?)")
                    set_task("error", f"#{new_nr}: Kein Bild – ChatGPT-Limit erreicht?", 0)
                    return
                append_log(f"✅ [5/8] Bild erstellt")

            # 6. Audio generieren
            _row_check = ir.find_row(new_nr, input_file)
            if _row_check and _row_check.get("status_audio") == "X":
                append_log(f"✅ [6/8] Audio bereits vorhanden – überspringe")
            else:
                set_task("running", f"[6/8] Audio #{new_nr} via ElevenLabs...", 50)
                append_log(f"⏳ [6/8] Audio via ElevenLabs generieren...")
                run_script_logged(["generate_audio.py", "--story", new_nr])
                append_log(f"✅ [6/8] Audio fertig")

            # 6.5 Bild auf Stereotyp-Passung prüfen (nur wenn Video noch nicht existiert)
            _row_check = ir.find_row(new_nr, input_file)
            if not (_row_check and _row_check.get("status_video") == "X"):
                set_task("running", f"[6.5/8] Bild auf Stereotyp-Passung prüfen...", 58)
                append_log(f"⏳ [6.5/8] Claude prüft ob Bild zum Stereotyp passt...")
                check_proc = subprocess.run(
                    [sys.executable, "check_image_match.py", "--story", new_nr],
                    capture_output=True, text=True, encoding="utf-8", errors="replace",
                    cwd=str(Path(__file__).parent)
                )
                check_reason = check_proc.stdout.strip() or check_proc.stderr.strip()
                if check_proc.returncode == 0:
                    append_log(f"✅ [6.5/8] Bild passt zum Stereotyp: {check_reason}")
                elif check_proc.returncode == 2:
                    append_log(f"❌ [6.5/8] Bild passt NICHT zum Stereotyp!")
                    append_log(f"   Grund: {check_reason}")
                    append_log(f"   ➡ Bitte Bild ersetzen und erneut 'Sofort Posten' starten.")
                    set_task("error", f"#{new_nr}: Bild passt nicht zum Stereotyp – bitte prüfen & ersetzen", 0)
                    return
                else:
                    append_log(f"⚠️  [6.5/8] Bildprüfung nicht möglich: {check_reason} – weiter")

            # 7. Video rendern
            _row_check = ir.find_row(new_nr, input_file)
            if _row_check and _row_check.get("status_video") == "X":
                append_log(f"✅ [7/8] Video bereits vorhanden – überspringe")
            else:
                set_task("running", f"[7/8] Video #{new_nr} rendern (ffmpeg + Untertitel)...", 63)
                append_log(f"⏳ [7/8] Video rendern mit Karaoke-Untertiteln...")
                run_script_logged(["generate_videos.py", "--story", new_nr, "--subtitles", "--bg-music", "Demeter - O.P..mp3", "--bg-music-volume", "0.2"])
                append_log(f"✅ [7/8] Video gerendert")

            # Git sync: sicherstellen dass CSV-Push vor dem Post erfolgreich war
            try:
                repo_root = str(Path(__file__).parent.parent)
                no_win = subprocess.CREATE_NO_WINDOW
                subprocess.run(["git", "pull", "--rebase"], cwd=repo_root,
                               capture_output=True, creationflags=no_win, timeout=30)
                subprocess.run(["git", "push"], cwd=repo_root,
                               capture_output=True, creationflags=no_win, timeout=30)
                append_log("✅ CSV-Push synchronisiert")
            except Exception as e:
                append_log(f"⚠️  Git-Sync fehlgeschlagen (wird trotzdem gepostet): {e}")

            # 8. Instagram + YouTube posten
            set_task("running", f"[8/8] Poste #{new_nr} auf Instagram + YouTube...", 80)
            append_log("─" * 40)
            append_log(f"⏳ [8/8] Poste auf Instagram + YouTube...")
            env = os.environ.copy()
            env["FORCE_POST"] = "1"
            env["STORY_NR"] = new_nr

            proc = subprocess.Popen(
                [sys.executable, "instagram_poster.py"],
                cwd=Path(__file__).parent,
                env=env,
                stdout=subprocess.PIPE,
                stderr=subprocess.STDOUT,
                text=True,
                encoding="utf-8",
                errors="replace",
                creationflags=subprocess.CREATE_NO_WINDOW,
            )
            for line in proc.stdout:
                line = line.rstrip()
                if line:
                    append_log(line)
            proc.wait()

            # Prüfe ob Post tatsächlich erfolgreich war (auch wenn returncode != 0)
            import input_reader as ir
            posted_row = ir.find_row(new_nr, "1_input/1_input_file.txt")
            post_success = posted_row and posted_row.get("insta_post", "").strip() == "X"

            if post_success:
                append_log(f"✅ [8/8] #{new_nr} auf Instagram + YouTube gepostet!")
                try:
                    ws.cell(pending_row_idx, status_col).value = "X"
                    wb.save(str(STORY_JETZT_XLSX))
                    append_log("✅ story_jetzt.xlsx: Status=X gesetzt")
                except Exception as e:
                    append_log(f"⚠️  xlsx aktualisieren fehlgeschlagen: {e}")
            else:
                append_log(f"⚠️  [8/8] Posten fehlgeschlagen (insta_post nicht gesetzt)")

            set_task("running", "Dashboard aktualisieren...", 97)
            refresh_dashboard()
            set_task("complete", f"#{new_nr} '{stereotyp}' vollständig gepostet!", 100)
        except Exception as e:
            import traceback
            set_task("error", f"{e}\n{traceback.format_exc()}", 0)

    set_task("running", "Starte Sofort-Posten...", 3, log=[])
    threading.Thread(target=task, daemon=True).start()
    return jsonify({"status": "started"})


if __name__ == "__main__":
    print("[+] Dashboard: http://localhost:5000")
    app.run(host="0.0.0.0", port=5000, debug=False)
